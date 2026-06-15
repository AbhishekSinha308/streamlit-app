import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import os
import ast
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

# ── File paths ────────────────────────────────────────────────────────────────
weather_db_path = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\weatherDB_data.xlsx"
query1_path     = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\query_results.xlsx"
query2_path     = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\query2_results.xlsx"
Data_validation = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\Data_validation.xlsx"

# ── Styling ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
NORM_FILL   = PatternFill("solid", fgColor="FFFFFF")
BSIDE       = Side(style="thin", color="B0C4D8")
CBORDER     = Border(left=BSIDE, right=BSIDE, top=BSIDE, bottom=BSIDE)
POS_FONT    = Font(color="006400", name="Arial", size=10)
NEG_FONT    = Font(color="8B0000", name="Arial", size=10)
STD_FONT    = Font(name="Arial", size=10)

# ── Also define green/red FILL for background highlight ──────────────────────
POS_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")   # light red

def write_sheet(ws, data):
    cols      = list(data.columns)
    diff_cols = {i + 1 for i, c in enumerate(cols)
                 if c in ('Difference', 'RE_Solar_Comparison')}
    for ci, col in enumerate(cols, 1):
        cell           = ws.cell(1, ci, col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = CBORDER
    for ri, row in enumerate(data.itertuples(index=False), 2):
        fill = ALT_FILL if ri % 2 == 0 else NORM_FILL
        for ci, val in enumerate(row, 1):
            v    = None if (isinstance(val, float) and math.isnan(val)) else val
            cell = ws.cell(ri, ci, v)
            cell.border    = CBORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if ci in diff_cols and v is not None and isinstance(v, (int, float)):
                try:
                    is_zero    = round(abs(float(v)), 2) == 0.00
                    cell.font  = POS_FONT if is_zero else NEG_FONT
                    cell.fill  = POS_FILL if is_zero else NEG_FILL
                except Exception:
                    cell.font  = STD_FONT
                    cell.fill  = fill
            else:
                cell.font = STD_FONT
                cell.fill = fill
    for ci, col in enumerate(cols, 1):
        max_w = max(
            len(str(col)),
            data.iloc[:, ci - 1].astype(str).str.len().max() if len(data) else 0
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 30)
    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = ws.dimensions


# =============================================================================
# STEP 1 -- Load & expand weatherDB
# =============================================================================
print("Loading weatherDB data...")
weather_data = pd.read_excel(weather_db_path)
weather_data = weather_data[['TIMESTAMP_DAY', 'LATITUDE', 'LONGITUDE', 'GHI', 'WIND_SPEED_925MB', 'FILE']]
weather_data['GHI']              = weather_data['GHI'].apply(
    lambda x: ast.literal_eval(x) if isinstance(x, str) else {})
weather_data['WIND_SPEED_925MB'] = weather_data['WIND_SPEED_925MB'].apply(
    lambda x: ast.literal_eval(x) if isinstance(x, str) else {})

expanded = []
for _, row in weather_data.iterrows():
    for time, ghi_val in row['GHI'].items():
        expanded.append([
            row['TIMESTAMP_DAY'], row['LATITUDE'], row['LONGITUDE'],
            int(time), ghi_val, row['WIND_SPEED_925MB'].get(time), row['FILE']
        ])

df_w = pd.DataFrame(
    expanded,
    columns=['TIMESTAMP_DAY', 'LATITUDE', 'LONGITUDE', 'TIME',
             'WeatherDB_GHI', 'WeatherDB_Wind_Speed_925MB', 'Weather_File']
)
df_w['LATITUDE']  = df_w['LATITUDE'].round(1)
df_w['LONGITUDE'] = df_w['LONGITUDE'].round(1)
df_w['_row']      = df_w.groupby(['LATITUDE', 'LONGITUDE']).cumcount()
print(f"  WeatherDB expanded rows: {len(df_w)}")


# =============================================================================
# STEP 2 -- Load GHI comparison data
# Confirmed col mapping: col5=RE_GHI, col6=HGT, col7=Weather_GHI
# =============================================================================
print("Loading GHI comparison data...")
wb1  = openpyxl.load_workbook(query1_path, data_only=True)
ws_g = wb1['Comparison Sheet']
ghi_rows = []
for r in range(2, ws_g.max_row + 1):
    lat = ws_g.cell(r, 1).value
    lon = ws_g.cell(r, 2).value
    if lat is None and lon is None:
        continue
    ghi_rows.append({
        'LATITUDE'   : round(float(lat), 1),
        'LONGITUDE'  : round(float(lon), 1),
        'HGT'        : ws_g.cell(r, 6).value,
        'Weather_GHI': ws_g.cell(r, 7).value,
        'RE_GHI'     : ws_g.cell(r, 5).value,
    })
df_ghi         = pd.DataFrame(ghi_rows)
df_ghi['_row'] = df_ghi.groupby(['LATITUDE', 'LONGITUDE']).cumcount()
ghi_latlon_set = set(zip(df_ghi['LATITUDE'], df_ghi['LONGITUDE']))
print(f"  GHI rows: {len(df_ghi)} | lat/lons: {sorted(ghi_latlon_set)}")


# =============================================================================
# STEP 3 -- Load Wind comparison data from Query2_Results
# Actual col mapping: col4=LAT, col5=LON, col7=wind_speed (RE_Wind_Speed)
# ⚠️ IMPORTANT: Wind data loaded from 6th block onwards (skip first 5 blocks)
# =============================================================================
print("Loading Wind comparison data...")
wb2         = openpyxl.load_workbook(query2_path, data_only=True)
sheet_names = wb2.sheetnames
print(f"  Sheets in query2: {sheet_names}")

# Auto-detect sheet: prefer 'Query2_Results', else fall back to first sheet
ws_w2 = wb2['Query2_Results'] if 'Query2_Results' in sheet_names else wb2[sheet_names[0]]

wind_rows = []
for r in range(2, ws_w2.max_row + 1):
    lat = ws_w2.cell(r, 4).value   # col D = LATITUDE
    lon = ws_w2.cell(r, 5).value   # col E = LONGITUDE
    if lat is None and lon is None:
        continue
    wind_rows.append({
        'LATITUDE'      : round(float(lat), 1),
        'LONGITUDE'     : round(float(lon), 1),
        'RE_Wind_Speed' : ws_w2.cell(r, 7).value,   # col G = wind_speed
    })

df_wind         = pd.DataFrame(wind_rows)
df_wind['_row'] = df_wind.groupby(['LATITUDE', 'LONGITUDE']).cumcount()

# ── Skip first 5 blocks (blocks 1-5), start from 6th block onwards ──────────
df_wind = df_wind[df_wind['_row'] >= 6].copy()
df_wind['_row'] = df_wind.groupby(['LATITUDE', 'LONGITUDE']).cumcount()

wind_latlon_set = set(zip(df_wind['LATITUDE'], df_wind['LONGITUDE']))
print(f"  Wind rows: {len(df_wind)} | lat/lons: {sorted(wind_latlon_set)}")


# =============================================================================
# STEP 4 -- Merge on LATITUDE + LONGITUDE + _row (positional)
# =============================================================================
print("Merging data...")
df = df_w.merge(df_ghi,  on=['LATITUDE', 'LONGITUDE', '_row'], how='left')
df = df.merge(df_wind,   on=['LATITUDE', 'LONGITUDE', '_row'], how='left')
df.drop(columns=['_row'], inplace=True)

# ── Difference: WeatherDB_Wind_Speed_925MB - RE_Wind_Speed ──────────────────
df['Difference'] = df.apply(
    lambda r: round(r['WeatherDB_Wind_Speed_925MB'] - r['RE_Wind_Speed'], 2)
              if (r['LATITUDE'], r['LONGITUDE']) in wind_latlon_set
                 and pd.notna(r['WeatherDB_Wind_Speed_925MB'])
                 and pd.notna(r['RE_Wind_Speed'])
              else None,
    axis=1
)

# ── RE_Solar_Comparison: WeatherDB_GHI - RE_GHI ─────────────────────────────
df['RE_Solar_Comparison'] = df.apply(
    lambda r: round(r['WeatherDB_GHI'] - r['RE_GHI'], 2)
              if (r['LATITUDE'], r['LONGITUDE']) in ghi_latlon_set
                 and pd.notna(r['WeatherDB_GHI']) and pd.notna(r['RE_GHI'])
              else None,
    axis=1
)

# ── Blank WeatherDB_GHI for lat/lons with no GHI comparison source ───────────
mask_no_ghi = df.apply(
    lambda r: (r['LATITUDE'], r['LONGITUDE']) not in ghi_latlon_set, axis=1)
df.loc[mask_no_ghi, 'WeatherDB_GHI'] = None

col_order = [
    'TIMESTAMP_DAY', 'LATITUDE', 'LONGITUDE', 'Weather_File', 'TIME',
    'WeatherDB_Wind_Speed_925MB', 'RE_Wind_Speed', 'Difference',
    'HGT', 'WeatherDB_GHI', 'Weather_GHI', 'RE_GHI', 'RE_Solar_Comparison'
]
df = df[[c for c in col_order if c in df.columns]]
print(f"  Merged shape: {df.shape}")


# =============================================================================
# STEP 5 -- Write workbook
# =============================================================================
if os.path.exists(Data_validation):
    os.remove(Data_validation)

wb = openpyxl.Workbook()
wb.remove(wb.active)

for (lat, lon), grp in df.groupby(['LATITUDE', 'LONGITUDE'], sort=False):
    sheet_name = f"Lat{lat}_Lon{lon}".replace('.', 'p')[:31]
    ws = wb.create_sheet(sheet_name)
    g  = grp.dropna(axis=1, how='all').reset_index(drop=True)
    write_sheet(ws, g)
    key   = (lat, lon)
    notes = []
    if key not in ghi_latlon_set:
        notes.append("no GHI source -> RE_GHI/RE_Solar_Comparison blank")
    if key not in wind_latlon_set:
        notes.append("no Wind source -> RE_Wind_Speed/Difference blank")
    # FIX: replaced emoji with plain text to avoid cp1252 encoding errors
    note_str = f" | [WARN] {'; '.join(notes)}" if notes else ""
    print(f"  Sheet '{sheet_name}': {len(g)} rows | cols: {list(g.columns)}{note_str}")

wb.save(Data_validation)
print(f"\nDone. {len(wb.sheetnames)} sheets written to:\n  {Data_validation}")
print("Sheets:", wb.sheetnames)