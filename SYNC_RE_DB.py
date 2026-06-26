

import pandas as pd
import ast
import sys
import traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from typing import List, Tuple, Dict, Any, Optional


sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")


CONFIG = {

    'weather_file': r'\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha\WeatherDB_data.xlsx',
    'query_file':   r'\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha\query_results.xlsx',
    'sync_output_file':  r'\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha\Sync_from_RE.xlsx',
    'mismatch_output_file': r'\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha\Mismatch_data.xlsx',

    'group1_coords': None,
    'group2_coords': None,

   
    'ghi_sheet':      'Query1_Results',
    'wind_sheet':     'Query2_Results',


    'decimal_places': 2,
    'header_color':   '4472C4',
    'group_color':    'D9E1F2',
    'column_width':   16,

    'skip_first_blocks_group1': 0,  
    'skip_first_blocks_group2': 6,   

    'mismatch_tolerance': 0.01,
}


def print_banner(text: str, width: int = 100) -> None:
    """Print formatted banner."""
    print("\n" + "=" * width)
    print(f" {text:^{width-2}} ")
    print("=" * width + "\n")

def print_step(step_num: int, total: int, text: str) -> None:
    """Print step indicator."""
    print(f"\n[STEP {step_num}/{total}] {text}\n")

def print_success(text: str) -> None:
    """Print success message."""
    print(f"[✓ OK] {text}")

def print_info(text: str) -> None:
    """Print info message."""
    print(f"[INFO] {text}")

def print_warn(text: str) -> None:
    """Print warning message."""
    print(f"[WARN] {text}")

def print_error(text: str) -> None:
    """Print error message."""
    print(f"[✗ ERROR] {text}")

def parse_dict_string(dict_str) -> dict:
    """Parse dictionary string from Excel cells."""
    try:
        return ast.literal_eval(dict_str)
    except (ValueError, SyntaxError):
        return {}

def validate_files(weather_file: str, query_file: str) -> bool:
    """Validate input files exist and are readable."""
    try:
        if not Path(weather_file).exists():
            print_error(f"File not found: {weather_file}")
            return False
        if not Path(query_file).exists():
            print_error(f"File not found: {query_file}")
            return False
        print_success("All input files found")
        return True
    except Exception as e:
        print_error(f"File validation error: {str(e)}")
        return False

def validate_output_path(output_file: str) -> bool:
    """Validate output path is writable."""
    try:
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        print_success(f"Output path writable: {output_path.parent}")
        return True
    except Exception as e:
        print_error(f"Output path error: {str(e)}")
        return False


def read_and_prepare_data(weather_file: str, query_file: str, ghi_sheet: str, wind_sheet: str):
    """Read and prepare data from all input files."""
    print_info("Reading input files...")
    
    try:
        weather_db = pd.read_excel(weather_file)
        query1     = pd.read_excel(query_file, sheet_name=ghi_sheet)
        query2     = pd.read_excel(query_file, sheet_name=wind_sheet)

        print(f"  WeatherDB: {len(weather_db)} rows")
        print(f"  Query1 ({ghi_sheet}): {len(query1)} rows")
        print(f"  Query2 ({wind_sheet}): {len(query2)} rows")

        # Parse dictionary columns
        weather_db['GHI_dict']  = weather_db['GHI'].apply(parse_dict_string)
        weather_db['WIND_dict'] = weather_db['WIND_SPEED_925MB'].apply(parse_dict_string)

        # Parse dates
        weather_db['date'] = pd.to_datetime(weather_db['TIMESTAMP_DAY']).dt.date
        query1['date']     = pd.to_datetime(query1['weather_date']).dt.date
        query2['date']     = pd.to_datetime(query2['weather_date']).dt.date

        print_success("Data prepared and parsed")
        return weather_db, query1, query2
        
    except Exception as e:
        print_error(f"Data reading error: {str(e)}")
        raise

def auto_detect_coordinates(query_df: pd.DataFrame, lat_col: str, lon_col: str) -> List[Tuple[float, float]]:
    """Auto-detect unique coordinate pairs from query results."""
    coords = query_df[[lat_col, lon_col]].drop_duplicates().values.tolist()
    coords = [(float(lat), float(lon)) for lat, lon in coords]
    print_info(f"Auto-detected {len(coords)} coordinate pairs:")
    for i, (lat, lon) in enumerate(coords, 1):
        print(f"  [{i}] Lat: {lat}, Lon: {lon}")
    return coords

def create_workbook_with_formatting(header_color: str, group_color: str):
    """Create and format Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sync_from_RE'
    
    styles = {
        'header_fill': PatternFill(start_color=header_color, end_color=header_color, fill_type='solid'),
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'group_fill':  PatternFill(start_color=group_color,  end_color=group_color,  fill_type='solid'),
        'group_font':  Font(bold=True, size=10),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
    }
    return wb, ws, styles

def add_group_header(ws, row: int, title: str, styles: dict) -> int:
    """Add group header to worksheet."""
    ws[f'A{row}'] = title
    ws[f'A{row}'].font = styles['group_font']
    ws[f'A{row}'].fill = styles['group_fill']
    return row + 1

def add_column_headers(ws, row: int, headers: list, styles: dict) -> int:
    """Add column headers to worksheet."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value     = header
        cell.font      = styles['header_font']
        cell.fill      = styles['header_fill']
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border    = styles['border']
    return row + 1

def set_column_widths(ws, num_columns: int, width: int) -> None:
    """Set column widths in worksheet."""
    for col in range(1, num_columns + 1):
        ws.column_dimensions[chr(64 + col)].width = width

def save_workbook(wb, output_file: str) -> bool:
    """Save workbook to file."""
    try:
        print_info(f"Saving workbook to {output_file}...")
        wb.save(output_file)
        print_success("File saved successfully!")
        return True
    except Exception as e:
        print_error(f"Error saving file: {str(e)}")
        return False

def process_sync_group(ws, row: int, weather_db: pd.DataFrame, query_df: pd.DataFrame, 
                       coords: List[Tuple[float, float]], styles: dict, 
                       group_name: str, weather_col_dict: dict, query_col_dict: dict, 
                       output_col_map: dict, decimal_places: int, formula_config: dict,
                       skip_first_blocks: int = 0) -> Tuple[int, int]:
    """
    Process and write comparison group data with optional data alignment.
    
    Args:
        ws: Worksheet object
        row: Current row number
        weather_db: Weather database dataframe
        query_df: Query results dataframe
        coords: List of (latitude, longitude) tuples
        styles: Formatting styles dictionary
        group_name: Name of group (e.g., 'GHI COMPARISON')
        weather_col_dict: Weather column mappings
        query_col_dict: Query column mappings
        output_col_map: Output column specifications
        decimal_places: Decimal precision for numbers
        formula_config: Formula configuration
        skip_first_blocks: Number of blocks to skip per day
    
    Returns:
        Tuple of (next_row, count_written)
    """
    print_info(f"Processing GROUP - {group_name}...")
    if skip_first_blocks > 0:
        print(f"  Data alignment: Query data SHIFTED UP")
        print(f"  Block 1 onwards uses query rows {skip_first_blocks+1}+ (shifted from row 7)")
    
    row = add_group_header(ws, row, f'GROUP - {group_name}', styles)
    headers = list(output_col_map.keys())
    row = add_column_headers(ws, row, headers, styles)
    
    count = 0
    errors = 0
    fmt = f'0.{"0" * decimal_places}'
    
    weather_col_letter = formula_config['weather_col']
    query_col_letter = formula_config['query_col']
    
    for lat, lon in coords:
        weather_rows = weather_db[
            (weather_db[weather_col_dict['latitude']] == lat) & 
            (weather_db[weather_col_dict['longitude']] == lon)
        ].sort_values('date')
        
        # Get query data
        query_rows = query_df[
            (query_df[query_col_dict['latitude']] == lat) & 
            (query_df[query_col_dict['longitude']] == lon)
        ]
        
        if weather_rows.empty:
            print_warn(f"No weather data for {lat}/{lon}")
            errors += 1
            continue
            
        if query_rows.empty:
            print_warn(f"No query data for {lat}/{lon}")
            errors += 1
            continue
        
        print(f"  Processing {lat}, {lon}: {len(weather_rows)} weather, {len(query_rows)} query rows")
        
        q_idx = skip_first_blocks
        
        for _, w_row in weather_rows.iterrows():
            data_dict = w_row[weather_col_dict['data_dict']]
            if not data_dict:
                continue
            
            for time_key, data_value in data_dict.items():
                if q_idx >= len(query_rows):
                    break
                
                try:
                    q_row = query_rows.iloc[q_idx]
                    
                    # Write each column
                    col_idx = 1
                    for header, source_info in output_col_map.items():
                        cell = ws.cell(row=row, column=col_idx)
                        value = None
                        
                        # Determine value source
                        if source_info['source'] == 'weather':
                            value = w_row[source_info['column']]
                        elif source_info['source'] == 'data_dict':
                            value = data_value if source_info['column'] == 'value' else time_key
                        elif source_info['source'] == 'query':
                            value = q_row[source_info['column']]
                        elif source_info['source'] == 'formula':
                            # Write formula
                            formula_str = f'={weather_col_letter}{row}-{query_col_letter}{row}'
                            cell.value = formula_str
                            cell.number_format = fmt
                            value = None
                        
                        # Set cell value
                        if source_info['source'] != 'formula' and value is not None:
                            if isinstance(value, (int, float)):
                                cell.value = value
                                if source_info.get('number_format'):
                                    cell.number_format = fmt
                            else:
                                cell.value = value
                        
                        col_idx += 1
                    
                    count += 1
                    q_idx += 1
                    row += 1
                    
                except Exception as e:
                    print_warn(f"Row {row}: {str(e)}")
                    errors += 1
                    q_idx += 1
                    row += 1
    
    print_success(f"{group_name}: {count} rows written")
    if errors:
        print_warn(f"Total errors: {errors}")
    
    return row + 2, count

def generate_sync_file(config: dict) -> bool:
    """
    Generate Sync_from_RE.xlsx file.
    
    Returns:
        Tuple of (success: bool, sync_df: pd.DataFrame)
    """
    print_banner("PHASE 1: GENERATING SYNC_FROM_RE.XLSX")
    
    try:
        print_step(1, 5, "Validating input files")
        if not validate_files(config['weather_file'], config['query_file']):
            return False, None
        
        if not validate_output_path(config['sync_output_file']):
            return False, None
        
        print_step(2, 5, "Reading and preparing data")
        weather_db, query1, query2 = read_and_prepare_data(
            config['weather_file'], config['query_file'],
            config['ghi_sheet'],    config['wind_sheet']
        )
        
        print_step(3, 5, "Auto-detecting coordinates")
        group1_coords = config['group1_coords'] or auto_detect_coordinates(
            query1, 'latitude_name', 'longitude_name'
        )
        group2_coords = config['group2_coords'] or auto_detect_coordinates(
            query2, 'latitude_name', 'longitude_name'
        )
        

        print_step(4, 5, "Creating Excel workbook")
        wb, ws, styles = create_workbook_with_formatting(
            config['header_color'], config['group_color']
        )
        print_success("Workbook created")
        
        row = 1
        
        print("\n" + "=" * 100)
        print("PROCESSING GROUP 1 - GHI COMPARISON (No Data Skip)")
        print("=" * 100)
        
        weather_col_dict_g1 = {
            'latitude': 'LATITUDE',
            'longitude': 'LONGITUDE',
            'data_dict': 'GHI_dict'
        }
        query_col_dict_g1 = {
            'latitude': 'latitude_name',
            'longitude': 'longitude_name'
        }
        output_col_map_g1 = {
            'LATITUDE': {'source': 'weather', 'column': 'LATITUDE'},
            'LONGITUDE': {'source': 'weather', 'column': 'LONGITUDE'},
            'TIMESTAMP_DAY': {'source': 'weather', 'column': 'TIMESTAMP_DAY'},
            'GHI_time_key': {'source': 'data_dict', 'column': 'time_key'},
            'GHI': {'source': 'data_dict', 'column': 'value', 'number_format': True},
            'latitude_name': {'source': 'query', 'column': 'latitude_name'},
            'longitude_name': {'source': 'query', 'column': 'longitude_name'},
            'weather_date': {'source': 'query', 'column': 'weather_date'},
            'weather_time': {'source': 'query', 'column': 'weather_time'},
            'process_file_name': {'source': 'query', 'column': 'process_file_name'},
            'param_name': {'source': 'query', 'column': 'param_name'},
            'param_value': {'source': 'query', 'column': 'param_value', 'number_format': True},
            'FILE': {'source': 'weather', 'column': 'FILE'},
            'GHI_Compare': {'source': 'formula'}
        }
        
        formula_config_g1 = {
            'weather_col': 'E',
            'query_col': 'L',
            'result_col': 'N'
        }
        
        row, count_g1 = process_sync_group(
            ws, row, weather_db, query1, group1_coords, styles,
            'GHI COMPARISON',
            weather_col_dict_g1, query_col_dict_g1, output_col_map_g1,
            config['decimal_places'],
            formula_config_g1,
            skip_first_blocks=config['skip_first_blocks_group1']
        )
        
        print(f"\nFormula: GHI (Column E) - param_value (Column L) = GHI_Compare (Column N)")

        print("\n" + "=" * 100)
        print("PROCESSING GROUP 2 - WIND SPEED COMPARISON (Data Skip Applied Per Day)")
        print("=" * 100)
        
        weather_col_dict_g2 = {
            'latitude': 'LATITUDE',
            'longitude': 'LONGITUDE',
            'data_dict': 'WIND_dict'
        }
        query_col_dict_g2 = {
            'latitude': 'latitude_name',
            'longitude': 'longitude_name'
        }
        output_col_map_g2 = {
            'LATITUDE': {'source': 'weather', 'column': 'LATITUDE'},
            'LONGITUDE': {'source': 'weather', 'column': 'LONGITUDE'},
            'TIMESTAMP_DAY': {'source': 'weather', 'column': 'TIMESTAMP_DAY'},
            'Wind_time_key': {'source': 'data_dict', 'column': 'time_key'},
            'WIND_SPEED_925MB': {'source': 'data_dict', 'column': 'value', 'number_format': True},
            'weather_date': {'source': 'query', 'column': 'weather_date'},
            'master_file_name': {'source': 'query', 'column': 'master_file_name'},
            'time': {'source': 'query', 'column': 'time'},
            'latitude_name': {'source': 'query', 'column': 'latitude_name'},
            'longitude_name': {'source': 'query', 'column': 'longitude_name'},
            'wind_speed': {'source': 'query', 'column': 'wind_speed', 'number_format': True},
            'level': {'source': 'query', 'column': 'level'},
            'FILE': {'source': 'weather', 'column': 'FILE'},
            'Wind_Compare': {'source': 'formula'}
        }
        
        formula_config_g2 = {
            'weather_col': 'E',
            'query_col': 'K',
            'result_col': 'N'
        }
        
        row, count_g2 = process_sync_group(
            ws, row, weather_db, query2, group2_coords, styles,
            'WIND SPEED COMPARISON',
            weather_col_dict_g2, query_col_dict_g2, output_col_map_g2,
            config['decimal_places'],
            formula_config_g2,
            skip_first_blocks=config['skip_first_blocks_group2']
        )
        
        print(f"\nFormula: WIND_SPEED_925MB (Column E) - wind_speed (Column K) = Wind_Compare (Column N)")
    
        print_step(5, 5, "Formatting columns")
        set_column_widths(ws, 14, config['column_width'])
        

        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        ws.conditional_formatting.add(
            f'N1:N{ws.max_row}',
            CellIsRule(operator='between', formula=['-0.01', '0.01'], fill=green_fill)
        )
        print_success("Column formatting applied")
        
  
        if not save_workbook(wb, config['sync_output_file']):
            return False, None
        
        print_info("Reading generated sync file for mismatch extraction...")
        sync_df = pd.read_excel(config['sync_output_file'], sheet_name=0, header=None)
        
        print_banner(f"[✓ PHASE 1 COMPLETE] {count_g1 + count_g2} rows written")
        print(f"  GHI COMPARISON: {count_g1} rows")
        print(f"  WIND COMPARISON: {count_g2} rows")
        print(f"  Output: {config['sync_output_file']}\n")
        
        return True, sync_df
        
    except Exception as e:
        print_error(f"Sync generation failed: {str(e)}")
        traceback.print_exc()
        return False, None


def extract_mismatches_from_dataframe(df: pd.DataFrame, tolerance: float) -> Tuple[list, list, list, list]:
    """
    Extract mismatch rows from sync dataframe.
    
    Returns:
        Tuple of (ghi_mismatches, wind_mismatches, ghi_headers, wind_headers)
    """
    print_info("Parsing sync file structure...")
    
    ghi_group_row = None
    wind_group_row = None
    ghi_header_row = None
    wind_header_row = None
    

    for idx, row in df.iterrows():
        first_cell = str(row[0]) if pd.notna(row[0]) else ""
        
        if "GROUP - GHI" in first_cell:
            ghi_group_row = idx
            ghi_header_row = idx + 1
            print_success(f"Found GROUP 1 (GHI) at row {idx}")
            
        if "GROUP - WIND" in first_cell:
            wind_group_row = idx
            wind_header_row = idx + 1
            print_success(f"Found GROUP 2 (WIND) at row {idx}")
    
    if ghi_group_row is None or wind_group_row is None:
        print_error("Could not find GROUP headers in file")
        return [], [], [], []
    
    print_info("Extracting GHI mismatches...")
    ghi_headers = df.iloc[ghi_header_row].tolist()
    ghi_data_start = ghi_header_row + 1
    ghi_data_end = wind_group_row
    ghi_rows = df.iloc[ghi_data_start:ghi_data_end].copy()
    ghi_rows.columns = ghi_headers
    
    ghi_col_idx = None
    param_value_col_idx = None
    
    for idx, col in enumerate(ghi_headers):
        if pd.notna(col):
            col_str = str(col).strip()
            if col_str == 'GHI':
                ghi_col_idx = idx
            elif col_str == 'param_value':
                param_value_col_idx = idx
    
    ghi_mismatches = []
    if ghi_col_idx is not None and param_value_col_idx is not None:
        for idx, row in ghi_rows.iterrows():
            try:
                ghi_val = float(row.iloc[ghi_col_idx]) if pd.notna(row.iloc[ghi_col_idx]) else 0
                param_val = float(row.iloc[param_value_col_idx]) if pd.notna(row.iloc[param_value_col_idx]) else 0
                comparison = ghi_val - param_val
                
                if abs(comparison) > tolerance:
                    ghi_mismatches.append((idx, row, comparison))
            except (ValueError, TypeError):
                pass
        
        print_success(f"Found {len(ghi_mismatches)} GHI mismatches (tolerance >{tolerance})")
    
    print_info("Extracting WIND mismatches...")
    wind_headers = df.iloc[wind_header_row].tolist()
    wind_data_start = wind_header_row + 1
    wind_rows = df.iloc[wind_data_start:].copy()
    wind_rows.columns = wind_headers
    
    wind_speed_col_idx = None
    wind_speed_query_col_idx = None
    
    for idx, col in enumerate(wind_headers):
        if pd.notna(col):
            col_str = str(col).strip()
            if col_str == 'WIND_SPEED_925MB':
                wind_speed_col_idx = idx
            elif col_str == 'wind_speed':
                wind_speed_query_col_idx = idx
    
    wind_mismatches = []
    if wind_speed_col_idx is not None and wind_speed_query_col_idx is not None:
        for idx, row in wind_rows.iterrows():
            try:
                wind_val = float(row.iloc[wind_speed_col_idx]) if pd.notna(row.iloc[wind_speed_col_idx]) else 0
                wind_query_val = float(row.iloc[wind_speed_query_col_idx]) if pd.notna(row.iloc[wind_speed_query_col_idx]) else 0
                comparison = wind_val - wind_query_val
                
                if abs(comparison) > tolerance:
                    wind_mismatches.append((idx, row, comparison))
            except (ValueError, TypeError):
                pass
        
        print_success(f"Found {len(wind_mismatches)} WIND mismatches (tolerance >{tolerance})")
    
    return ghi_mismatches, wind_mismatches, ghi_headers, wind_headers

def write_mismatches_to_excel(output_file: str, ghi_mismatches: list, wind_mismatches: list, 
                              ghi_headers: list, wind_headers: list, 
                              header_color: str, group_color: str) -> Tuple[int, int]:
    """Write mismatches to Excel file."""
    print_info(f"Writing mismatches to {output_file}...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mismatch_Data'
    
    styles = {
        'header_fill': PatternFill(start_color=header_color, end_color=header_color, fill_type='solid'),
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'group_fill': PatternFill(start_color=group_color, end_color=group_color, fill_type='solid'),
        'group_font': Font(bold=True, size=10),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    }
    
    row = 1
    total_written = 0
    
    if ghi_mismatches:
        print_info(f"Writing {len(ghi_mismatches)} GHI mismatch rows...")
        
        ws[f'A{row}'] = 'GROUP - GHI COMPARISON'
        ws[f'A{row}'].font = styles['group_font']
        ws[f'A{row}'].fill = styles['group_fill']
        row += 1
        

        for col_idx, header in enumerate(ghi_headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = styles['border']
        
        cell = ws.cell(row=row, column=len(ghi_headers) + 1)
        cell.value = 'GHI_Compare'
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        row += 1
 
        for orig_idx, data_row, comparison in ghi_mismatches:
            for col_idx, header in enumerate(ghi_headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                value = data_row.iloc[col_idx - 1]
                cell.value = value
                cell.border = styles['border']
                if header in ['GHI', 'param_value']:
                    cell.number_format = '0.00'
            
            cell = ws.cell(row=row, column=len(ghi_headers) + 1)
            cell.value = comparison
            cell.number_format = '0.00'
            cell.border = styles['border']
            
            row += 1
            total_written += 1
        
        row += 2
        print_success(f"Wrote {len(ghi_mismatches)} GHI mismatch rows")
    if wind_mismatches:
        print_info(f"Writing {len(wind_mismatches)} WIND mismatch rows...")
        
        ws[f'A{row}'] = 'GROUP - WIND SPEED COMPARISON'
        ws[f'A{row}'].font = styles['group_font']
        ws[f'A{row}'].fill = styles['group_fill']
        row += 1
        
        for col_idx, header in enumerate(wind_headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = styles['border']
        
        cell = ws.cell(row=row, column=len(wind_headers) + 1)
        cell.value = 'Wind_Compare'
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        row += 1
        
        # Data rows
        for orig_idx, data_row, comparison in wind_mismatches:
            for col_idx, header in enumerate(wind_headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                value = data_row.iloc[col_idx - 1]
                cell.value = value
                cell.border = styles['border']
                if header in ['WIND_SPEED_925MB', 'wind_speed']:
                    cell.number_format = '0.00'
            
            cell = ws.cell(row=row, column=len(wind_headers) + 1)
            cell.value = comparison
            cell.number_format = '0.00'
            cell.border = styles['border']
            
            row += 1
            total_written += 1
        
        print_success(f"Wrote {len(wind_mismatches)} WIND mismatch rows")
    
    max_cols = max(len(ghi_headers) + 1, len(wind_headers) + 1) if (ghi_headers or wind_headers) else 1
    for col in range(1, max_cols + 1):
        ws.column_dimensions[chr(64 + col)].width = 16
    
    try:
        wb.save(output_file)
        print_success("Mismatch file saved successfully!")
        return len(ghi_mismatches), len(wind_mismatches)
    except Exception as e:
        print_error(f"Error saving mismatch file: {str(e)}")
        return 0, 0

def extract_mismatches(config: dict, sync_df: pd.DataFrame) -> bool:
    """
    Extract mismatches from sync file.
    
    Returns:
        bool: True if successful
    """
    print_banner("PHASE 2: EXTRACTING MISMATCHES")
    
    try:
        if sync_df is None:
            print_error("No sync dataframe provided")
            return False
        
        print_step(1, 2, "Parsing sync file and extracting mismatches")
        ghi_mismatches, wind_mismatches, ghi_headers, wind_headers = extract_mismatches_from_dataframe(
            sync_df, config['mismatch_tolerance']
        )
        
        print_step(2, 2, "Writing mismatches to Excel file")
        ghi_count, wind_count = write_mismatches_to_excel(
            config['mismatch_output_file'],
            ghi_mismatches, wind_mismatches,
            ghi_headers, wind_headers,
            config['header_color'], config['group_color']
        )
        
        print_banner(f"[✓ PHASE 2 COMPLETE] {ghi_count + wind_count} mismatch rows extracted")
        print(f"  GHI Mismatches: {ghi_count} rows")
        print(f"  WIND Mismatches: {wind_count} rows")
        print(f"  Output: {config['mismatch_output_file']}\n")
        
        return True
        
    except Exception as e:
        print_error(f"Mismatch extraction failed: {str(e)}")
        traceback.print_exc()
        return False


def main() -> int:
    """Main orchestration function."""
    
    start_time = datetime.now()
    print_banner("COMPLETE SYNC & MISMATCH EXTRACTOR - ALL IN ONE")
    print(f"Start time: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    try:
        success, sync_df = generate_sync_file(CONFIG)
        if not success or sync_df is None:
            print_error("Sync generation failed")
            return 1
        
        success = extract_mismatches(CONFIG, sync_df)
        if not success:
            print_error("Mismatch extraction failed")
            return 2
        

        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        print_banner("[✓ SUCCESS] ALL OPERATIONS COMPLETED")
        print(f"{'EXECUTION SUMMARY':^100}")
        print(f"{'-' * 100}")
        print(f"  Phase 1: Sync_from_RE.xlsx generated successfully")
        print(f"  Phase 2: Mismatch_data.xlsx extracted successfully")
        print(f"\n{'OUTPUT FILES':^100}")
        print(f"{'-' * 100}")
        print(f"  1. {CONFIG['sync_output_file']}")
        print(f"  2. {CONFIG['mismatch_output_file']}")
        print(f"\n{'TIMING':^100}")
        print(f"{'-' * 100}")
        print(f"  Start: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  End:   {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"  Duration: {duration:.2f} seconds")
        print("=" * 100 + "\n")
        
        return 0
        
    except Exception as e:
        print_error(f"Unexpected error: {str(e)}")
        traceback.print_exc()
        return 3


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)