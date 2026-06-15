"""
File_Monitoring(ECM10).py - XLSX OUTPUT ONLY VERSION

This script:
✅ Works through Streamlit UI
✅ Works when run individually  
✅ Outputs colors in both cases
✅ Saves to correct OUTPUT_DIR (network share)
✅ Generates XLSX file only (no CSV)
✅ Includes professional styling
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import os
import glob
from datetime import datetime
from pathlib import Path

# ============================================================================
# ✅ Get output directory from environment variable (set by Streamlit)
# ============================================================================
DEFAULT_OUTPUT_DIR = r"\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha"

OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", DEFAULT_OUTPUT_DIR))
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================================
# ✅ Color codes for terminal output
# ============================================================================
class Colors:
    GREEN = '\033[92m'      # Success
    RED = '\033[91m'        # Error
    YELLOW = '\033[93m'     # Warning
    BLUE = '\033[94m'       # Info
    CYAN = '\033[96m'       # Process
    MAGENTA = '\033[95m'    # Status
    RESET = '\033[0m'       # Reset color

# ============================================================================
# ✅ Logging functions
# ============================================================================
def log(message, color=Colors.RESET):
    """Print with color codes and immediate flush."""
    output = f"{color}{message}{Colors.RESET}"
    print(output, flush=True)
    sys.stdout.flush()

def log_section(title):
    """Log a section header."""
    log("\n" + "=" * 80, Colors.CYAN)
    log(f"  {title}", Colors.CYAN)
    log("=" * 80, Colors.CYAN)

def log_step(step_num, total, description):
    """Log a step in the process."""
    log(f"\n[{step_num}/{total}] {description}", Colors.YELLOW)

# Try to import netCDF4 for reading .nc files
try:
    import netCDF4 as nc
    HAS_NETCDF4 = True
except ImportError:
    HAS_NETCDF4 = False

# Fallback: try scipy
try:
    from scipy.io import netcdf_file
    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

log(f"📦 netCDF4 available : {HAS_NETCDF4}", Colors.BLUE)
log(f"📦 scipy   available : {HAS_SCIPY}", Colors.BLUE)

# ============================================================================
# Input file paths (network share)
# ============================================================================
share_path        = r"\\172.16.0.65\Share\24_23_QA_Team\Abhishek_Sinha"
ftp_path          = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\FileHistory.csv"
latlon_path       = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\LatLonCounts.csv"
weather_path      = r"\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\Weather_Updated_LatLong.xlsx"

# Output path - uses OUTPUT_DIR
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
output_xlsx_path = OUTPUT_DIR / "File_Monitoring.xlsx"

log_section("FILE MONITORING PROCESS - STARTED")
log(f"📁 Output Directory: {OUTPUT_DIR}", Colors.BLUE)
log(f"📁 Output File: {output_xlsx_path.name}", Colors.BLUE)

# ── Excel Styling ──────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
NORM_FILL     = PatternFill("solid", fgColor="FFFFFF")
BSIDE         = Side(style="thin", color="B0C4D8")
CBORDER       = Border(left=BSIDE, right=BSIDE, top=BSIDE, bottom=BSIDE)
STD_FONT      = Font(name="Arial", size=10)
POS_FILL      = PatternFill("solid", fgColor="C6EFCE")
NEG_FILL      = PatternFill("solid", fgColor="FFC7CE")
POS_FONT      = Font(color="006400", name="Arial", size=10)
NEG_FONT      = Font(color="8B0000", name="Arial", size=10)
FALLBACK_FILL = PatternFill("solid", fgColor="FFF2CC")
FALLBACK_FONT = Font(name="Arial", size=10, italic=True, color="7B5A00")


# =============================================================================
# HELPER -- Read latitude & longitude counts from a .nc file
# =============================================================================
def get_nc_lat_lon_counts(nc_file_path):
    """
    Opens a NetCDF file and returns (lat_count, lon_count).
    Tries netCDF4 first, then scipy as fallback.
    Returns (None, None) if file can't be read.
    """
    lat_names = ['latitude', 'lat', 'LAT', 'LATITUDE', 'y', 'Y']
    lon_names = ['longitude', 'lon', 'LON', 'LONGITUDE', 'x', 'X']

    if HAS_NETCDF4:
        try:
            with nc.Dataset(nc_file_path, 'r') as ds:
                lat_count = None
                lon_count = None
                for name in lat_names:
                    if name in ds.variables:
                        lat_count = len(ds.variables[name][:])
                        break
                for name in lon_names:
                    if name in ds.variables:
                        lon_count = len(ds.variables[name][:])
                        break
                # fallback: check dimensions if variables not found
                if lat_count is None:
                    for name in lat_names:
                        if name in ds.dimensions:
                            lat_count = ds.dimensions[name].size
                            break
                if lon_count is None:
                    for name in lon_names:
                        if name in ds.dimensions:
                            lon_count = ds.dimensions[name].size
                            break
                return lat_count, lon_count
        except Exception as e:
            log(f"    [netCDF4 error] {nc_file_path}: {e}", Colors.RED)

    if HAS_SCIPY:
        try:
            with netcdf_file(nc_file_path, 'r', mmap=False) as ds:
                lat_count = None
                lon_count = None
                for name in lat_names:
                    if name in ds.variables:
                        lat_count = len(ds.variables[name].data)
                        break
                for name in lon_names:
                    if name in ds.variables:
                        lon_count = len(ds.variables[name].data)
                        break
                return lat_count, lon_count
        except Exception as e:
            log(f"    [scipy error] {nc_file_path}: {e}", Colors.RED)

    return None, None


# =============================================================================
# STEP 1 -- Scan share for all .nc files
# =============================================================================
try:
    log_step(1, 6, f"Scanning share for .nc files: {share_path}")
    nc_files_on_share = {}

    all_nc = glob.glob(os.path.join(share_path, "*.nc"))
    log(f"   Found {len(all_nc)} .nc files on share", Colors.GREEN)

    for i, nc_path in enumerate(all_nc, 1):
        fname   = os.path.basename(nc_path)
        mod_ts  = os.path.getmtime(nc_path)
        ftp_dt  = datetime.fromtimestamp(mod_ts).strftime('%d-%m-%Y %H:%M')

        log(f"   [{i}/{len(all_nc)}] Reading {fname} ...", Colors.BLUE)
        lat_count, lon_count = get_nc_lat_lon_counts(nc_path)
        log(f"       ✅ Lat: {lat_count}, Lon: {lon_count}", Colors.GREEN)

        nc_files_on_share[fname] = {
            'full_path'   : nc_path,
            'FTP Datetime': ftp_dt,
            'NC_FileLatitude Count' : lat_count,
            'NC_FileLongitude Count': lon_count,
        }

except Exception as e:
    log(f"   ❌ Could not scan share: {e}", Colors.RED)
    nc_files_on_share = {}


# =============================================================================
# STEP 2 -- Load CSVs
# =============================================================================
try:
    log_step(2, 6, "Loading input CSV files")
    
    log(f"   Loading FileHistory from:\n      {ftp_path}", Colors.BLUE)
    file_history = pd.read_csv(ftp_path)
    log(f"   ✅ FileHistory: {len(file_history)} rows, {len(file_history.columns)} cols", Colors.GREEN)
    
    log(f"   Loading LatLonCounts from:\n      {latlon_path}", Colors.BLUE)
    latlon = pd.read_csv(latlon_path)
    log(f"   ✅ LatLonCounts: {len(latlon)} rows, {len(latlon.columns)} cols", Colors.GREEN)
    
    log(f"   Loading Weather data from:\n      {weather_path}", Colors.BLUE)
    weather_data = pd.read_excel(weather_path, sheet_name='Sheet2')
    log(f"   ✅ Weather data: {len(weather_data)} rows", Colors.GREEN)

    # LatLonCounts lookup (FTP registered files — highest priority)
    ll_lookup = latlon.set_index('FTP File Name').to_dict(orient='index')
    log(f"   📊 FTP files registered in LatLonCounts: {len(ll_lookup)}", Colors.MAGENTA)

except Exception as e:
    log(f"   ❌ Error loading CSV files: {e}", Colors.RED)
    raise


# =============================================================================
# STEP 3 -- Build output rows
# =============================================================================
log_step(3, 6, "Building output rows with priority lookup")

rows           = []
fallback_flags = []
latlon_count   = 0
share_count    = 0
missing_count  = 0

for idx, (_, fh_row) in enumerate(file_history.iterrows(), 1):
    if idx % 100 == 0:
        log(f"   Processing row {idx}/{len(file_history)}...", Colors.BLUE)

    db_fname  = str(fh_row.get('DB_File Name', '')).strip()
    ftp_fname = str(fh_row.get('FTP File Name', '')).strip()

    # Priority 1: LatLonCounts
    lookup_key = ftp_fname if ftp_fname and ftp_fname in ll_lookup else (
                 db_fname  if db_fname  and db_fname  in ll_lookup else None)

    if lookup_key:
        ll_data     = ll_lookup[lookup_key]
        out_ftp     = lookup_key
        out_ftp_dt  = fh_row.get('FTP Datetime') or ll_data.get('FTP Datetime')
        out_lat     = ll_data.get('NC_FileLatitude Count')
        out_lon     = ll_data.get('NC_FileLongitude Count')
        is_fallback = False
        latlon_count += 1

    # Priority 2: .nc file on share
    elif db_fname and db_fname in nc_files_on_share:
        share_data  = nc_files_on_share[db_fname]
        out_ftp     = db_fname
        out_ftp_dt  = share_data['FTP Datetime']
        out_lat     = share_data['NC_FileLatitude Count']
        out_lon     = share_data['NC_FileLongitude Count']
        is_fallback = True
        share_count += 1

    # Priority 3: not found
    else:
        out_ftp     = db_fname or ftp_fname or ''
        out_ftp_dt  = fh_row.get('FTP Datetime', '')
        out_lat     = None
        out_lon     = None
        is_fallback = True
        missing_count += 1
        if missing_count <= 5:
            log(f"   ⚠️  '{db_fname}' not found in LatLonCounts or share", Colors.YELLOW)
        elif missing_count == 6:
            log(f"   ... and {missing_count - 5} more files not found", Colors.YELLOW)

    row = {
        'FTP File Name'          : out_ftp,
        'FTP Datetime'           : out_ftp_dt,
        'NC_FileLatitude Count'  : out_lat,
        'NC_FileLongitude Count' : out_lon,
    }

    # Add all remaining FileHistory columns
    skip = {'FTP File Name', 'FTP Datetime', 'NC_FileLatitude Count', 'NC_FileLongitude Count'}
    for col in file_history.columns:
        if col not in skip:
            row[col] = fh_row[col]

    rows.append(row)
    fallback_flags.append(is_fallback)

combined_df     = pd.DataFrame(rows)
fallback_series = pd.Series(fallback_flags, index=combined_df.index)

log(f"\n   ✅ Total rows processed: {len(combined_df)}", Colors.GREEN)
log(f"   ✅ From LatLonCounts   : {latlon_count}", Colors.GREEN)
log(f"   ✅ From share .nc files: {share_count}", Colors.YELLOW)
log(f"   ⚠️  Missing/not found   : {missing_count}", Colors.RED if missing_count > 0 else Colors.GREEN)


# =============================================================================
# STEP 4 -- Computed columns
# =============================================================================
log_step(4, 6, "Computing derived columns")

try:
    combined_df['Difference LatLong'] = (
        pd.to_numeric(combined_df['NC_FileLatitude Count'],  errors='coerce') -
        pd.to_numeric(combined_df['NC_FileLongitude Count'], errors='coerce')
    )
    log(f"   ✅ Computed 'Difference LatLong' column", Colors.GREEN)

    if 'File Processed End Hours' in combined_df.columns and 'File Processed Start Hours' in combined_df.columns:
        combined_df['Difference Time'] = (
            pd.to_numeric(combined_df['File Processed End Hours'],   errors='coerce') -
            pd.to_numeric(combined_df['File Processed Start Hours'], errors='coerce')
        )
        log(f"   ✅ Computed 'Difference Time' column", Colors.GREEN)

    if 'File Name' in weather_data.columns and 'Count' in weather_data.columns:
        weather_lookup = weather_data.set_index('File Name')['Count'].to_dict()
        combined_df['Total Lat&Long Data processed to DB'] = combined_df['FTP File Name'].map(weather_lookup)
        log(f"   ✅ Mapped weather data counts", Colors.GREEN)

except Exception as e:
    log(f"   ⚠️  Error computing columns: {e}", Colors.YELLOW)


# =============================================================================
# STEP 5 -- Column ordering
# =============================================================================
log_step(5, 6, "Organizing columns")

priority_cols = [
    'FTP File Name', 'FTP Datetime',
    'NC_FileLatitude Count', 'NC_FileLongitude Count',
    'DB_File Name', 'File Processed into DB at',
    'Total Lat&Long Data processed to DB',
    'File Status', 'Description',
    'File Processed Start Hours', 'File Processed End Hours',
    'Difference LatLong', 'Difference Time'
]
remaining   = [c for c in combined_df.columns if c not in priority_cols]
final_cols  = [c for c in priority_cols if c in combined_df.columns] + remaining
combined_df = combined_df[final_cols]

log(f"   ✅ Final column count: {len(final_cols)}", Colors.GREEN)
log(f"      Columns: {', '.join(final_cols[:5])}...", Colors.BLUE)


# =============================================================================
# STEP 6 -- Save styled Excel file
# =============================================================================
log_step(6, 6, "Saving styled XLSX file")

try:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File_Monitoring"

    diff_col_names = {'Difference LatLong', 'Difference Time'}
    cols     = list(combined_df.columns)
    diff_idx = {i + 1 for i, c in enumerate(cols) if c in diff_col_names}

    # Header row
    for ci, col in enumerate(cols, 1):
        cell           = ws.cell(1, ci, col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = CBORDER
    ws.row_dimensions[1].height = 32

    # Data rows with styling
    for ri, (row, is_fallback) in enumerate(zip(combined_df.itertuples(index=False), fallback_series), 2):
        default_fill = FALLBACK_FILL if is_fallback else (ALT_FILL if ri % 2 == 0 else NORM_FILL)
        default_font = FALLBACK_FONT if is_fallback else STD_FONT

        for ci, val in enumerate(row, 1):
            v = None if (isinstance(val, float) and math.isnan(val)) else val
            cell = ws.cell(ri, ci, v)
            cell.border    = CBORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if ci in diff_idx and v is not None and isinstance(v, (int, float)):
                try:
                    is_zero   = round(abs(float(v)), 2) == 0.00
                    cell.font = POS_FONT if is_zero else NEG_FONT
                    cell.fill = POS_FILL if is_zero else NEG_FILL
                except Exception:
                    cell.font = default_font
                    cell.fill = default_fill
            else:
                cell.font = default_font
                cell.fill = default_fill

    # Auto column widths
    for ci, col in enumerate(cols, 1):
        max_w = max(
            len(str(col)),
            combined_df.iloc[:, ci - 1].astype(str).str.len().max() if len(combined_df) else 0
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 40)

    # Freeze panes and auto filter
    ws.freeze_panes    = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Save file
    wb.save(output_xlsx_path)
    log(f"   ✅ Saved XLSX: {output_xlsx_path.name}", Colors.GREEN)

    # Print summary
    log(f"\n{'─' * 80}", Colors.CYAN)
    log(f"📊 SUMMARY:", Colors.MAGENTA)
    log(f"   Total records processed: {len(combined_df)}", Colors.BLUE)
    log(f"   From LatLonCounts: {latlon_count} ({100*latlon_count/len(combined_df):.1f}%)", Colors.GREEN)
    log(f"   From share .nc: {share_count} ({100*share_count/len(combined_df):.1f}%)", Colors.YELLOW)
    log(f"   Missing: {missing_count} ({100*missing_count/len(combined_df):.1f}%)", Colors.RED if missing_count > 0 else Colors.GREEN)
    log(f"\n📁 Output file:", Colors.MAGENTA)
    log(f"   {output_xlsx_path}", Colors.BLUE)
    
    # Print sample data
    log(f"\n📋 Sample output (first 3 rows):", Colors.MAGENTA)
    sample = combined_df.head(3).to_string()
    for line in sample.split('\n'):
        log(f"   {line}", Colors.BLUE)

except Exception as e:
    log(f"\n❌ Error saving file: {e}", Colors.RED)
    import traceback
    log(traceback.format_exc(), Colors.RED)
    sys.exit(1)


# =============================================================================
# COMPLETION
# =============================================================================
log_section("FILE MONITORING PROCESS - COMPLETED SUCCESSFULLY")
log(f"✅ Output file saved to:\n   {output_xlsx_path}", Colors.GREEN)
sys.exit(0)