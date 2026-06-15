import openpyxl
import json

def copy_data(file_path, weather_file_path):
  
    wb = openpyxl.load_workbook(file_path)
    
    source_sheet = wb["Query1_Results"]
    
    if "Comparison Sheet" in wb.sheetnames:
        target_sheet = wb["Comparison Sheet"]
    else:
        target_sheet = wb.create_sheet("Comparison Sheet")
  
    headers = {cell.value: cell.column for cell in source_sheet[1] if cell.value}
    required_columns = ["latitude_name", "longitude_name", "process_file_name", "param_name", "param_value"]

    col_indices = [headers[col] for col in required_columns if col in headers]

    target_sheet.delete_rows(1, target_sheet.max_row)

    target_sheet.append(required_columns)  
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        target_sheet.append([row[i - 1] for i in col_indices])
 
    weather_wb = openpyxl.load_workbook(weather_file_path)
    weather_sheet = weather_wb.active  
    
    weather_headers = {cell.value: cell.column for cell in weather_sheet[1] if cell.value}
    ghi_column_index = weather_headers.get("GHI")
    lat_index = weather_headers.get("LATITUDE")
    lon_index = weather_headers.get("LONGITUDE")
    timestamp_index = weather_headers.get("TIMESTAMP_DAY")
    
    if ghi_column_index and lat_index and lon_index and timestamp_index:
       
        target_sheet.cell(row=1, column=6, value="HGT")  
        target_sheet.cell(row=1, column=7, value="GHI")     
        allowed_locations = [(20.1, 74), (15.8, 78), (8.9, 77.8)]
    
        ghi_data_dict = {loc: [] for loc in allowed_locations}
    
        for row in weather_sheet.iter_rows(min_row=2, values_only=True):
            latitude = row[lat_index - 1]
            longitude = row[lon_index - 1]
            ghi_data = row[ghi_column_index - 1]
            
            if (latitude, longitude) in allowed_locations and isinstance(ghi_data, str):
                ghi_dict = json.loads(ghi_data.replace("'", '"')) 
                for key, value in ghi_dict.items():
                    ghi_data_dict[(latitude, longitude)].append((key, value))
     
        ghi_data_start_row = 2 
        for location in allowed_locations:
            for hgt, ghi in ghi_data_dict[location]:
                target_sheet.cell(row=ghi_data_start_row, column=6, value=hgt)
                target_sheet.cell(row=ghi_data_start_row, column=7, value=ghi)
                ghi_data_start_row += 1
        
        target_sheet.cell(row=1, column=8, value="Comparison (GHI - param_value)")

        for row in target_sheet.iter_rows(min_row=2, max_row=target_sheet.max_row, values_only=False):
            param_value_cell = row[4] 
            ghi_cell = row[6]  
            comparison_cell = row[7] 
            
            if param_value_cell.value is not None and ghi_cell.value is not None:
               
                comparison_cell.value = round(ghi_cell.value - param_value_cell.value, 2)
    
    wb.save(file_path)
    print("Data copied and comparison calculated successfully!")
    
file_path = r'\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\query_results.xlsx'
weather_file_path = r'\\172.16.0.65\share\24_23_QA_Team\Abhishek_Sinha\WeatherDB_data.xlsx'
copy_data(file_path, weather_file_path)
